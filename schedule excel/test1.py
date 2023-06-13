import openpyxl

def generate_excel_file(data):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Set column headers
    sheet['A1'] = 'Team Member'
    sheet['B1'] = 'Working Days'
    sheet['C1'] = 'Days Off'
    sheet['D1'] = 'Total Required'

    # Write data to the worksheet
    for row, member in enumerate(data, start=2):
        sheet.cell(row=row, column=1).value = member['name']
        sheet.cell(row=row, column=2).value = member['working_days']
        sheet.cell(row=row, column=3).value = member['days_off']
        sheet.cell(row=row, column=4).value = member['total_required']

    # Save the workbook
    workbook.save('team_schedule.xlsx')

# Example data
team_data = [
    {'name': 'John', 'working_days': 20, 'days_off': 5, 'total_required': 25},
    {'name': 'Jane', 'working_days': 22, 'days_off': 3, 'total_required': 25},
    {'name': 'Mike', 'working_days': 18, 'days_off': 7, 'total_required': 25}
]

# Generate the Excel file
generate_excel_file(team_data)
